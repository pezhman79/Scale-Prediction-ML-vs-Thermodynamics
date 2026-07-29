# -*- coding: utf-8 -*-
"""
Enhanced Mineral Scale Prediction Pipeline — RESULTS-SECTION VERSION
=============================================================================
Changes vs. previous version (per author's request for the paper's Results
section):

  1) Added Extra Trees as a 5th ML model (full Optuna tuning, same as the
     other tree-based models).
  2) "Thermodynamic (Corrected)" REMOVED from the final results table and
     from every comparison figure. Only "Thermodynamic" is kept as
     the physics-informed baseline in the final comparison.
     -> NOTE: the activity/Davies-correction machinery (ionic strength,
        Davies gamma, temperature-dependent Ksp for calcite/barite/celestite)
        is kept in the code, since it may still be needed for the Methods
        section narrative — it is simply excluded from the RESULTS outputs.
  3) Output figures restructured to exactly match the requested set:
        A) Results table  -> Accuracy/Precision/Recall/Specificity/F1/MCC/AUC
                              for Thermodynamic + 5 ML models
        B) Combined confusion-matrix figure (all 6 models, one panel each)
        C) Resampling effect:
              C1) grouped bar chart, ADASYN vs No-ADASYN (all 5 ML models)
              C2) confusion-matrix comparison: BEST model, with vs without
                  ADASYN
        D) Combined ROC-AUC curve (all 6 models)
        E) Permutation importance for the 5 ML models (one combined figure)
        F) SHAP analysis for the single BEST ML model only

  Removed (not requested for this results section, code kept simple):
     - radar chart, per-model individual bar charts, individual ROC curves,
       all PR curves, SI/SR naive-vs-corrected distribution figure. Say the
       word and any of these can be added back.

  4) NEW — everything stays in this ONE file (no separate script):
     After training/tuning/evaluating once, section (G) saves every ML
     model (with & without ADASYN) plus the test data to ./saved_models/.
     Section (H), immediately after, DELETES the in-memory models/results,
     RELOADS everything back from disk, and reruns every single result and
     figure (A, A2, B, C1, C2, D, E, F) a second time on the test set —
     using only the reloaded models. This proves the saved artifacts alone
     are enough to reproduce the paper's figures, and gives you a ready-made
     block to re-run any time you just want to regenerate figures without
     repeating Optuna tuning.

  5) THERMODYNAMIC MODEL — SI-ONLY VERSION (per author's request):
     The "naive" thermodynamic baseline (the one actually used in every
     results table/figure below) now computes the Saturation Index (SI)
     UNIFORMLY for all three minerals — calcite, barite, and celestite —
     using the same logarithmic definition:

         SI = log10(IAP / Ksp)

     Previously, barite and celestite were scored with a linear Saturation
     Ratio (SR = IAP/Ksp) while calcite used the logarithmic SI, an
     inconsistency in scale between minerals. barite_SR_naive() and
     celestite_SR_naive() have been replaced with barite_SI_naive() and
     celestite_SI_naive(), both returning log10(IAP/Ksp), so all three
     minerals are directly comparable and share one physical threshold
     (SI = 0). The binary decision rule and the continuous risk score used
     for ROC-AUC have been updated to match (see Section 6 and Section 8
     below). The "corrected" (activity/Davies-based) functions in Section 7
     are left untouched, kept only for the Methods-section narrative and
     not used in any reported result.

  6) HALITE (NaCl) AND GYPSUM (CaSO4) ADDED (per author's request):
     The thermodynamic baseline now covers FIVE scale-forming minerals
     instead of three: calcite (CaCO3), barite (BaSO4), celestite (SrSO4),
     halite (NaCl), and gypsum (CaSO4). Both new minerals follow the exact
     same SI = log10(IAP/Ksp) convention as the other three, in both the
     "naive" (Section 6, fixed Ksp constants, actually used in the results)
     and the "corrected" (Section 7, Davies activity coefficients + a
     temperature-dependent Ksp, Methods-narrative only) implementations.
     The binary decision (Section 8) and the continuous risk score are
     extended to the OR-logic / max-sigmoid across all five minerals.
     NOTE: the Davies equation is strictly valid only up to moderate ionic
     strength (I ~ 0.5 mol/kg); halite-saturated brines routinely exceed
     this range, so the "corrected" halite SI should be interpreted with
     added caution and is, in any case, not used in the reported results.
     The temperature-dependent Ksp expressions for halite and gypsum used
     in Section 7 are simplified empirical fits intended to preserve the
     same functional style as the existing barite/celestite correlations,
     not literature-calibrated equations of state; this is an acknowledged
     simplification, consistent with the pressure-independence limitation
     already noted for the other three minerals.

Requires: pip install shap   (in addition to the original dependencies)

Author: pejma (enhanced, results-section revision)
=============================================================================
"""

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.ticker import MaxNLocator
import seaborn as sns
import optuna

from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.base import clone
from sklearn.metrics import (classification_report, confusion_matrix, roc_auc_score,
                              f1_score, roc_curve, accuracy_score,
                              matthews_corrcoef, precision_score, recall_score)
from sklearn.preprocessing import RobustScaler
from sklearn.ensemble import RandomForestClassifier, ExtraTreesClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.inspection import permutation_importance
from xgboost import XGBClassifier
from lightgbm import LGBMClassifier

from imblearn.pipeline import Pipeline as ImbPipeline
from imblearn.over_sampling import ADASYN

import shap
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# GLOBAL STYLE  (journal-ready figure formatting)
# ---------------------------------------------------------------------------
plt.rcParams.update({
    "font.family":        "DejaVu Sans",
    "font.size":          10,
    "axes.titlesize":     10,
    "axes.labelsize":     10,
    "xtick.labelsize":    9,
    "ytick.labelsize":    9,
    "legend.fontsize":    8,
    "legend.frameon":     True,
    "legend.framealpha":  0.85,
    "legend.edgecolor":   "0.7",
    "axes.linewidth":     0.8,
    "axes.spines.top":    False,
    "axes.spines.right":  False,
    "xtick.direction":    "out",
    "ytick.direction":    "out",
    "xtick.major.size":   3.5,
    "ytick.major.size":   3.5,
    "figure.dpi":         150,
    "savefig.dpi":        300,
    "savefig.bbox":       "tight",
    "axes.grid":          True,
    "grid.color":         "0.88",
    "grid.linewidth":     0.5,
    "grid.linestyle":     "--",
})

# ---------------------------------------------------------------------------
# PALETTE  (colorblind-safe, consistent across all figures) — now 5 ML models
# ---------------------------------------------------------------------------
MODEL_NAMES  = ['Random Forest', 'Extra Trees', 'XGBoost', 'LightGBM', 'MLP']
MODEL_LABELS = MODEL_NAMES
COLORS       = ['#D55E00', '#E69F00', '#0072B2', '#009E73', '#999999']
MARKERS      = ['s', 'D', 'o', '^', 'X']
MODEL_COLOR  = dict(zip(MODEL_NAMES, COLORS))
MODEL_MARKER = dict(zip(MODEL_NAMES, MARKERS))
THERMO_COLOR_NAIVE = '#B2182B'   # only "Naive" survives in the results section

# ---------------------------------------------------------------------------
# EXCEL EXPORT HELPER — classification-report workbook, paper-ready formatting
#   Sheet 1  "Summary"           -> final_comparison (all 7 metrics, all models)
#   Sheet 2+ one per model       -> full classification report (per-class
#                                   Precision/Recall/F1/Support + Accuracy)
# ---------------------------------------------------------------------------

def export_classification_reports_excel(model_preds, y_true_arr, final_comparison_df,
                                          save_path, colors_map):
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    title_font = Font(bold=True, name='Calibri', size=13)
    cell_font = Font(name='Calibri', size=10.5)
    bold_cell_font = Font(name='Calibri', size=10.5, bold=True)

    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = 'Final Results Summary — Thermodynamic Model vs. ML Models'
    ws['A1'].font = title_font
    ws.merge_cells(f'A1:{chr(65 + len(final_comparison_df.columns) - 1)}1')

    for c, col_name in enumerate(final_comparison_df.columns, start=1):
        cell = ws.cell(row=3, column=c, value=col_name)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='404040', end_color='404040')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for r, (_, row) in enumerate(final_comparison_df.iterrows(), start=4):
        model_name = row['Model']
        hex_color = colors_map.get(model_name, '#7F7F7F').lstrip('#')
        for c, col_name in enumerate(final_comparison_df.columns, start=1):
            val = row[col_name]
            val = round(val, 4) if isinstance(val, (int, float, np.floating)) else val
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = bold_cell_font if c == 1 else cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')
            if c == 1:
                cell.fill = PatternFill('solid', start_color=hex_color, end_color=hex_color)
                cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10.5)

    ws.column_dimensions['A'].width = 26
    for col in list('BCDEFGH')[:len(final_comparison_df.columns) - 1]:
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = 'A4'

    # --- One classification-report sheet per model ---
    row_labels = ['No Scale', 'Scale', 'accuracy', 'macro avg', 'weighted avg']
    for name, y_pred in model_preds.items():
        report = classification_report(y_true_arr, y_pred, target_names=['No Scale', 'Scale'],
                                        output_dict=True)
        sheet_name = name.replace('(', '').replace(')', '')[:31]
        sh = wb.create_sheet(sheet_name)

        sh['A1'] = f'Classification Report — {name}'
        sh['A1'].font = title_font
        sh.merge_cells('A1:E1')

        headers = ['Class', 'Precision', 'Recall', 'F1-score', 'Support']
        hex_color = colors_map.get(name, '#404040').lstrip('#')
        for c, h in enumerate(headers, start=1):
            cell = sh.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = PatternFill('solid', start_color=hex_color, end_color=hex_color)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        r = 4
        for label in row_labels:
            if label == 'accuracy':
                sh.cell(row=r, column=1, value='Accuracy').font = bold_cell_font
                sh.cell(row=r, column=4, value=round(report['accuracy'], 4)).font = cell_font
                sh.cell(row=r, column=5, value=int(report['weighted avg']['support'])).font = cell_font
            else:
                d = report[label]
                sh.cell(row=r, column=1, value=label).font = bold_cell_font
                sh.cell(row=r, column=2, value=round(d['precision'], 4)).font = cell_font
                sh.cell(row=r, column=3, value=round(d['recall'], 4)).font = cell_font
                sh.cell(row=r, column=4, value=round(d['f1-score'], 4)).font = cell_font
                sh.cell(row=r, column=5, value=int(d['support'])).font = cell_font
            for c in range(1, 6):
                sh.cell(row=r, column=c).border = border
                if c > 1:
                    sh.cell(row=r, column=c).alignment = Alignment(horizontal='center')
            r += 1

        for col, width in zip('ABCDE', [16, 12, 12, 12, 10]):
            sh.column_dimensions[col].width = width

    wb.save(save_path)
    print(f"Classification-report Excel workbook saved to: {save_path}")

# =============================================================================
# 1) LOAD DATA
# =============================================================================

DATA_PATH = r"C:\python\scale_PT\scaledata_1.xlsx"
df = pd.read_excel(DATA_PATH)
df = (df
      .dropna()
      .drop_duplicates()
      .drop(columns=['Scale Type', 'Well No'], errors='ignore'))

TARGET_COL = 'Inspection Result'
if TARGET_COL not in df.columns:
    raise ValueError(f"Target column '{TARGET_COL}' not found in data!")

X = df.drop(columns=[TARGET_COL])
y = df[TARGET_COL]

classes = np.unique(y)
if len(classes) != 2:
    raise ValueError(f"Expected binary target; found {len(classes)} classes: {classes}")

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.25, random_state=5, stratify=y
)

cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

# =============================================================================
# 2) COLUMN MAPPING  (exact names as provided by the user)
# =============================================================================

def find_col(possible_names, columns):
    for p in possible_names:
        for c in columns:
            if p.lower().replace(" ", "") in c.lower().replace(" ", ""):
                return c
    return None

col_T    = find_col(["T"], X.columns)
col_P    = find_col(["P"], X.columns)
col_pH   = find_col(["pH"], X.columns)
col_Ca   = find_col(["Ca2+", "Ca"], X.columns)
col_Na   = find_col(["Na+", "Na"], X.columns)
col_Mg   = find_col(["Mg2+", "Mg"], X.columns)
col_Fe   = find_col(["Fe2+", "Fe"], X.columns)
col_HCO3 = find_col(["HCO30", "HCO3"], X.columns)
col_SO4  = find_col(["SO4"], X.columns)
col_Cl   = find_col(["Cl-", "Cl"], X.columns)
col_CO3  = find_col(["CO320", "CO3"], X.columns)
col_Ba   = find_col(["Ba2+", "Ba"], X.columns)
col_Sr   = find_col(["Sr2+", "Sr"], X.columns)
col_TDS  = find_col(["TDS"], X.columns)

print("Detected columns:")
for name, val in [("T", col_T), ("P", col_P), ("pH", col_pH), ("Ca", col_Ca),
                   ("Na", col_Na), ("Mg", col_Mg), ("Fe", col_Fe), ("HCO3", col_HCO3),
                   ("SO4", col_SO4), ("Cl", col_Cl), ("CO3", col_CO3), ("Ba", col_Ba),
                   ("Sr", col_Sr), ("TDS", col_TDS)]:
    print(f"  {name:6s}: {val}")

MW = {
    'Ca': 40.078, 'Na': 22.990, 'Mg': 24.305, 'Fe': 55.845,
    'HCO3': 61.016, 'SO4': 96.060, 'Cl': 35.453, 'CO3': 60.008,
    'Ba': 137.327, 'Sr': 87.620,
}

# =============================================================================
# 3) IONIC STRENGTH  (I = 0.5 * sum(ci * zi^2))
# =============================================================================

def get_molar(row, col, mw):
    if col is None or pd.isna(row.get(col, np.nan)):
        return 0.0
    return max(row[col], 0.0) / (mw * 1000.0)

def ionic_strength(row):
    ions = [
        (get_molar(row, col_Ca, MW['Ca']), 2),
        (get_molar(row, col_Na, MW['Na']), 1),
        (get_molar(row, col_Mg, MW['Mg']), 2),
        (get_molar(row, col_Fe, MW['Fe']), 2),
        (get_molar(row, col_HCO3, MW['HCO3']), 1),
        (get_molar(row, col_SO4, MW['SO4']), 2),
        (get_molar(row, col_Cl, MW['Cl']), 1),
        (get_molar(row, col_CO3, MW['CO3']), 2),
        (get_molar(row, col_Ba, MW['Ba']), 2),
        (get_molar(row, col_Sr, MW['Sr']), 2),
    ]
    I = 0.5 * sum(c * (z ** 2) for c, z in ions)
    return max(I, 1e-8)

# =============================================================================
# 4) DAVIES ACTIVITY COEFFICIENT  (kept for Methods section — not used in
#    the final RESULTS comparison anymore, since "Corrected" was removed
#    from the reported outputs)
# =============================================================================

def debye_huckel_A(T_K):
    T_C = T_K - 273.15
    return 0.4918 + 6.0435e-4 * T_C + 1.3132e-6 * T_C ** 2

def activity_coefficient_davies(z, I, T_K):
    A = debye_huckel_A(T_K)
    sqI = np.sqrt(I)
    log_gamma = -A * (z ** 2) * (sqI / (1 + sqI) - 0.3 * I)
    return 10 ** log_gamma

def to_kelvin(T_F):
    return (T_F - 32) * 5 / 9 + 273.15

# =============================================================================
# 5) TEMPERATURE-DEPENDENT SOLUBILITY PRODUCTS (kept for Methods section)
# =============================================================================

def logKsp_calcite(T_K):
    return (-171.9065 - 0.077993 * T_K + 2839.319 / T_K + 71.595 * np.log10(T_K))

def logKsp_barite(T_K):
    T_C = T_K - 273.15
    return -9.97 - 0.00028 * T_C + 0.0000068 * (T_C ** 2)

def logKsp_celestite(T_K):
    T_C = T_K - 273.15
    return -6.63 - 0.0022 * T_C + 0.0000091 * (T_C ** 2)

def logKsp_halite(T_K):
    """Simplified empirical fit (same functional style as barite/celestite
    above): NaCl solubility increases mildly with temperature, so logKsp
    rises slightly with T_C relative to the 25 degC reference (logKsp=1.57).
    Not a literature-calibrated equation of state — see module docstring
    item 6."""
    T_C = T_K - 273.15
    return 1.57 + 0.00069 * (T_C - 25)

def logKsp_gypsum(T_K):
    """Simplified empirical fit (same functional style as barite/celestite
    above): gypsum solubility first rises then falls with temperature,
    approximated here with a downward-curving quadratic relative to the
    25 degC reference (logKsp=-4.50). Not a literature-calibrated equation
    of state — see module docstring item 6."""
    T_C = T_K - 273.15
    return -4.50 - 0.0011 * (T_C - 25) - 0.0000068 * (T_C - 25) ** 2

def estimate_carbonate_from_ph(hco3_mol, ph):
    hco3_mol = max(hco3_mol, 1e-10)
    ratio = 10 ** (ph - 10.3)
    return max(hco3_mol * ratio, 1e-10)

# =============================================================================
# 6) NAIVE THERMODYNAMIC MODEL — SI computed UNIFORMLY for calcite, barite,
#    and celestite. This IS the model used in the final results below.
#
#    SI = log10( IAP / Ksp )
#
#    All three minerals share the same logarithmic definition, so a single
#    physical threshold (SI = 0) applies consistently across the board.
#    Ksp values here are the simple, non-activity-corrected constants
#    (no Davies correction, no full P,T-dependence) — hence "naive" — but
#    they are still all expressed on the same log(IAP/Ksp) scale.
# =============================================================================

def calcite_SI_naive(row):
    """SI(CaCO3) = log10(IAP) - logKsp, with a mildly temperature-adjusted logKsp."""
    try:
        Ca_mol = get_molar(row, col_Ca, MW['Ca'])
        HCO3_mol = get_molar(row, col_HCO3, MW['HCO3'])
        pH = row[col_pH]
        T_K = to_kelvin(row[col_T])
        CO3_mol = estimate_carbonate_from_ph(HCO3_mol, pH)
        IAP = max(Ca_mol, 1e-10) * CO3_mol
        logKsp = -8.48 + 0.01 * ((298.15 - T_K) / 10)
        return np.log10(IAP) - logKsp
    except Exception:
        return np.nan

def barite_SI_naive(row):
    """SI(BaSO4) = log10(IAP / Ksp), Ksp = 1e-10 (naive, non-activity-corrected)."""
    try:
        Ba_mol = get_molar(row, col_Ba, MW['Ba'])
        SO4_mol = get_molar(row, col_SO4, MW['SO4'])
        IAP = max(Ba_mol, 1e-10) * max(SO4_mol, 1e-10)
        Ksp = 1e-10
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def celestite_SI_naive(row):
    """SI(SrSO4) = log10(IAP / Ksp), Ksp = 10^-6.63 (naive, non-activity-corrected)."""
    try:
        Sr_mol = get_molar(row, col_Sr, MW['Sr'])
        SO4_mol = get_molar(row, col_SO4, MW['SO4'])
        IAP = max(Sr_mol, 1e-10) * max(SO4_mol, 1e-10)
        Ksp = 10 ** -6.63
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def halite_SI_naive(row):
    """SI(NaCl) = log10(IAP / Ksp), Ksp = 37.7 (naive, non-activity-corrected,
    25 degC reference value for Na+ x Cl- in mol^2/L^2)."""
    try:
        Na_mol = get_molar(row, col_Na, MW['Na'])
        Cl_mol = get_molar(row, col_Cl, MW['Cl'])
        IAP = max(Na_mol, 1e-10) * max(Cl_mol, 1e-10)
        Ksp = 37.7
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def gypsum_SI_naive(row):
    """SI(CaSO4.2H2O) = log10(IAP / Ksp), Ksp = 3.14e-5 (naive, non-activity-
    corrected, 25 degC reference value for Ca2+ x SO4^2- in mol^2/L^2)."""
    try:
        Ca_mol = get_molar(row, col_Ca, MW['Ca'])
        SO4_mol = get_molar(row, col_SO4, MW['SO4'])
        IAP = max(Ca_mol, 1e-10) * max(SO4_mol, 1e-10)
        Ksp = 3.14e-5
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

# =============================================================================
# 7) CORRECTED (activity-based) SI — kept for Methods narrative only.
#    NOT used anywhere in the results table/figures below. All three
#    minerals already share the same log10(IAP/Ksp) SI definition here.
# =============================================================================

def calcite_SI_corrected(row):
    try:
        T_K = to_kelvin(row[col_T])
        pH = row[col_pH]
        I = ionic_strength(row)
        Ca_mol = get_molar(row, col_Ca, MW['Ca'])
        HCO3_mol = get_molar(row, col_HCO3, MW['HCO3'])
        CO3_mol = estimate_carbonate_from_ph(HCO3_mol, pH)
        gamma_Ca = activity_coefficient_davies(2, I, T_K)
        gamma_CO3 = activity_coefficient_davies(2, I, T_K)
        IAP = (gamma_Ca * max(Ca_mol, 1e-10)) * (gamma_CO3 * CO3_mol)
        Ksp = 10 ** logKsp_calcite(T_K)
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def barite_SI_corrected(row):
    try:
        T_K = to_kelvin(row[col_T])
        I = ionic_strength(row)
        Ba_mol = get_molar(row, col_Ba, MW['Ba'])
        SO4_mol = get_molar(row, col_SO4, MW['SO4'])
        gamma_Ba = activity_coefficient_davies(2, I, T_K)
        gamma_SO4 = activity_coefficient_davies(2, I, T_K)
        IAP = (gamma_Ba * max(Ba_mol, 1e-10)) * (gamma_SO4 * max(SO4_mol, 1e-10))
        Ksp = 10 ** logKsp_barite(T_K)
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def celestite_SI_corrected(row):
    try:
        T_K = to_kelvin(row[col_T])
        I = ionic_strength(row)
        Sr_mol = get_molar(row, col_Sr, MW['Sr'])
        SO4_mol = get_molar(row, col_SO4, MW['SO4'])
        gamma_Sr = activity_coefficient_davies(2, I, T_K)
        gamma_SO4 = activity_coefficient_davies(2, I, T_K)
        IAP = (gamma_Sr * max(Sr_mol, 1e-10)) * (gamma_SO4 * max(SO4_mol, 1e-10))
        Ksp = 10 ** logKsp_celestite(T_K)
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def halite_SI_corrected(row):
    """SI(NaCl) = log10( [gamma_Na * Na_mol] * [gamma_Cl * Cl_mol] / Ksp ).
    NOTE: the Davies equation is only strictly valid up to I ~ 0.5 mol/kg;
    halite-saturated brines frequently exceed this range (see module
    docstring item 6), so this value should be treated as approximate."""
    try:
        T_K = to_kelvin(row[col_T])
        I = ionic_strength(row)
        Na_mol = get_molar(row, col_Na, MW['Na'])
        Cl_mol = get_molar(row, col_Cl, MW['Cl'])
        gamma_Na = activity_coefficient_davies(1, I, T_K)
        gamma_Cl = activity_coefficient_davies(1, I, T_K)
        IAP = (gamma_Na * max(Na_mol, 1e-10)) * (gamma_Cl * max(Cl_mol, 1e-10))
        Ksp = 10 ** logKsp_halite(T_K)
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

def gypsum_SI_corrected(row):
    """SI(CaSO4.2H2O) = log10( [gamma_Ca * Ca_mol] * [gamma_SO4 * SO4_mol] / Ksp )."""
    try:
        T_K = to_kelvin(row[col_T])
        I = ionic_strength(row)
        Ca_mol = get_molar(row, col_Ca, MW['Ca'])
        SO4_mol = get_molar(row, col_SO4, MW['SO4'])
        gamma_Ca = activity_coefficient_davies(2, I, T_K)
        gamma_SO4 = activity_coefficient_davies(2, I, T_K)
        IAP = (gamma_Ca * max(Ca_mol, 1e-10)) * (gamma_SO4 * max(SO4_mol, 1e-10))
        Ksp = 10 ** logKsp_gypsum(T_K)
        return np.log10(IAP / Ksp)
    except Exception:
        return np.nan

# =============================================================================
# 8) APPLY NAIVE THERMODYNAMIC MODEL TO TEST SET  (SI-only, all 5 minerals)
# =============================================================================

thermo_df = X_test.copy()
thermo_df["I"]                  = thermo_df.apply(ionic_strength, axis=1)
thermo_df["Calcite_SI_naive"]   = thermo_df.apply(calcite_SI_naive, axis=1)
thermo_df["Barite_SI_naive"]    = thermo_df.apply(barite_SI_naive, axis=1)
thermo_df["Celestite_SI_naive"] = thermo_df.apply(celestite_SI_naive, axis=1)
thermo_df["Halite_SI_naive"]    = thermo_df.apply(halite_SI_naive, axis=1)
thermo_df["Gypsum_SI_naive"]    = thermo_df.apply(gypsum_SI_naive, axis=1)

SI_COLS_NAIVE = ["Calcite_SI_naive", "Barite_SI_naive", "Celestite_SI_naive",
                  "Halite_SI_naive", "Gypsum_SI_naive"]

def thermo_decision(df_, si_cols, threshold_si=0.0):
    """Scale-forming if SI > 0 for ANY of the five minerals (OR-logic),
    all five evaluated on the same log10(IAP/Ksp) scale."""
    combined = (df_[si_cols[0]] > threshold_si).astype(int)
    for c in si_cols[1:]:
        combined = combined | (df_[c] > threshold_si).astype(int)
    return combined

y_pred_thermo_naive = thermo_decision(thermo_df, SI_COLS_NAIVE).values
y_true = y_test.values

def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-x))

def thermo_risk_score(df_, si_cols):
    """Continuous risk score for ROC-AUC: sigmoid(SI) per mineral, anchored
    at the physical equilibrium point SI = 0, with the overall score taken
    as the maximum across the five minerals (mirrors the OR-logic used for
    the binary decision above)."""
    risks = [sigmoid(df_[c].values) for c in si_cols]
    return np.maximum.reduce(risks)

proba_thermo_naive = thermo_risk_score(thermo_df, SI_COLS_NAIVE)

def safe_auc(y_true_, proba_):
    try:
        return roc_auc_score(y_true_, proba_)
    except Exception:
        return np.nan

def compute_thermo_metrics(y_te, y_pred, y_proba):
    cm = confusion_matrix(y_te, y_pred)
    tn, fp, fn, tp = cm.ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan
    return {
        'Accuracy':    accuracy_score(y_te, y_pred),
        'Precision':   precision_score(y_te, y_pred, zero_division=0),
        'Recall':      recall_score(y_te, y_pred, zero_division=0),
        'Specificity': specificity,
        'F1':          f1_score(y_te, y_pred, average='weighted'),
        'MCC':         matthews_corrcoef(y_te, y_pred),
        'ROC-AUC':     safe_auc(y_te, y_proba),
        'cm':          cm,
    }

metrics_thermo_naive = compute_thermo_metrics(y_true, y_pred_thermo_naive, proba_thermo_naive)

print("\n" + "=" * 80)
print("Thermodynamic Model (SI-only, calcite + barite + celestite + halite + gypsum)")
print("=" * 80)
print(f"Thermodynamic -> Acc: {metrics_thermo_naive['Accuracy']:.4f} | "
      f"Prec: {metrics_thermo_naive['Precision']:.4f} | Rec: {metrics_thermo_naive['Recall']:.4f} | "
      f"Spec: {metrics_thermo_naive['Specificity']:.4f} | F1: {metrics_thermo_naive['F1']:.4f} | "
      f"MCC: {metrics_thermo_naive['MCC']:.4f} | AUC: {metrics_thermo_naive['ROC-AUC']:.4f}")
print("\n--- Classification Report: Thermodynamic Model ---")
print(classification_report(y_true, y_pred_thermo_naive, target_names=['No Scale', 'Scale']))

# =============================================================================
# 9) ML PIPELINES  (RF, Extra Trees, XGB, LightGBM, MLP)
# =============================================================================

def build_pipeline(estimator, use_adasyn=True):
    """
    IMPORTANT: we clone() the estimator here. Without this, calling
    build_pipeline(est, ...) twice with the SAME `est` object (once for the
    ADASYN pipeline, once for the no-ADASYN pipeline) would make both
    Pipelines hold a reference to the exact same classifier instance.
    Fitting one later (e.g. the no-ADASYN version) would then silently
    overwrite the fitted weights of the other (in-place), corrupting the
    already-trained ADASYN model. clone() gives each pipeline its own
    independent, unfitted copy with the same hyperparameters, so fitting
    one never touches the other.
    """
    steps = [('scaler', RobustScaler())]
    if use_adasyn:
        steps.append(('sampler', ADASYN(sampling_strategy='minority', random_state=42)))
    steps.append(('clf', clone(estimator)))
    return ImbPipeline(steps=steps)

def objective_rf(trial):
    params = {
        'n_estimators': trial.suggest_int('n_estimators', 100, 500),
        'max_depth': trial.suggest_int('max_depth', 5, 20),
        'min_samples_split': trial.suggest_int('min_samples_split', 2, 20),
        'min_samples_leaf': trial.suggest_int('min_samples_leaf', 1, 10),
        'max_features': trial.suggest_categorical('max_features', ['sqrt', 'log2', None]),
    }
    model = build_pipeline(RandomForestClassifier(random_state=42, **params))
    return float(np.mean(cross_val_score(model, X_train, y_train, cv=cv, scoring='roc_auc', n_jobs=-1)))

def objective_et(trial):
    params = {
        'n_estimators': trial.suggest_int('n_estimators', 100, 500),
        'max_depth': trial.suggest_int('max_depth', 5, 20),
        'min_samples_split': trial.suggest_int('min_samples_split', 2, 20),
        'min_samples_leaf': trial.suggest_int('min_samples_leaf', 1, 10),
        'max_features': trial.suggest_categorical('max_features', ['sqrt', 'log2', None]),
    }
    model = build_pipeline(ExtraTreesClassifier(random_state=42, **params))
    return float(np.mean(cross_val_score(model, X_train, y_train, cv=cv, scoring='roc_auc', n_jobs=-1)))

def objective_xgb(trial):
    params = {
        'n_estimators': trial.suggest_int('n_estimators', 100, 500),
        'max_depth': trial.suggest_int('max_depth', 3, 15),
        'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3),
        'subsample': trial.suggest_float('subsample', 0.6, 1.0),
        'colsample_bytree': trial.suggest_float('colsample_bytree', 0.6, 1.0),
        'min_child_weight': trial.suggest_int('min_child_weight', 1, 10),
        'gamma': trial.suggest_float('gamma', 0, 5),
        'reg_alpha': trial.suggest_float('reg_alpha', 0, 1),
        'reg_lambda': trial.suggest_float('reg_lambda', 0, 1),
    }
    model = build_pipeline(XGBClassifier(random_state=42, eval_metric='logloss', **params))
    return float(np.mean(cross_val_score(model, X_train, y_train, cv=cv, scoring='roc_auc', n_jobs=-1)))

def objective_lgbm(trial):
    params = {
        'n_estimators': trial.suggest_int('n_estimators', 100, 500),
        'max_depth': trial.suggest_int('max_depth', 3, 15),
        'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3),
        'num_leaves': trial.suggest_int('num_leaves', 15, 127),
        'subsample': trial.suggest_float('subsample', 0.6, 1.0),
        'colsample_bytree': trial.suggest_float('colsample_bytree', 0.6, 1.0),
        'reg_alpha': trial.suggest_float('reg_alpha', 0, 1),
        'reg_lambda': trial.suggest_float('reg_lambda', 0, 1),
    }
    model = build_pipeline(LGBMClassifier(random_state=42, verbose=-1, **params))
    return float(np.mean(cross_val_score(model, X_train, y_train, cv=cv, scoring='roc_auc', n_jobs=-1)))

def objective_mlp(trial):
    n_layers = trial.suggest_int('n_layers', 1, 3)
    layer_size = trial.suggest_categorical('layer_size', [32, 64, 128])
    hidden = tuple([layer_size] * n_layers)
    params = {
        'hidden_layer_sizes': hidden,
        'alpha': trial.suggest_float('alpha', 1e-5, 1e-1, log=True),
        'learning_rate_init': trial.suggest_float('learning_rate_init', 1e-4, 1e-2, log=True),
        'max_iter': 1000,
        'early_stopping': True,
    }
    model = build_pipeline(MLPClassifier(random_state=42, **params))
    return float(np.mean(cross_val_score(model, X_train, y_train, cv=cv, scoring='roc_auc', n_jobs=-1)))

N_TRIALS = 40

studies = {}
objectives = {
    'Random Forest': objective_rf,
    'Extra Trees':   objective_et,
    'XGBoost':       objective_xgb,
    'LightGBM':      objective_lgbm,
    'MLP':           objective_mlp,
}

for name, obj in objectives.items():
    print("\n" + "=" * 70)
    print(f"{name.upper()} HYPERPARAMETER OPTIMIZATION")
    print("=" * 70)
    study = optuna.create_study(direction='maximize', sampler=optuna.samplers.TPESampler(seed=42))
    study.optimize(obj, n_trials=N_TRIALS, show_progress_bar=True)
    studies[name] = study
    print(f"Best {name} params: {study.best_params}")
    print(f"Best ROC-AUC: {study.best_value:.4f}")

# -----------------------------------------------------------------------
# Build final estimators with best params
# -----------------------------------------------------------------------

best_rf = RandomForestClassifier(random_state=42, **studies['Random Forest'].best_params)
best_et = ExtraTreesClassifier(random_state=42, **studies['Extra Trees'].best_params)
best_xgb = XGBClassifier(random_state=42, eval_metric='logloss', **studies['XGBoost'].best_params)
best_lgbm = LGBMClassifier(random_state=42, verbose=-1, **studies['LightGBM'].best_params)

mlp_params = studies['MLP'].best_params.copy()
n_layers = mlp_params.pop('n_layers')
layer_size = mlp_params.pop('layer_size')
mlp_params['hidden_layer_sizes'] = tuple([layer_size] * n_layers)
mlp_params['max_iter'] = 1000
mlp_params['early_stopping'] = True
best_mlp = MLPClassifier(random_state=42, **mlp_params)

base_estimators = {
    'Random Forest': best_rf,
    'Extra Trees':   best_et,
    'XGBoost':       best_xgb,
    'LightGBM':      best_lgbm,
    'MLP':           best_mlp,
}

pipelines = {name: build_pipeline(est, use_adasyn=True) for name, est in base_estimators.items()}

# =============================================================================
# 10) EVALUATION METRICS
# =============================================================================

def compute_full_metrics(y_te, y_pred, y_proba):
    cm = confusion_matrix(y_te, y_pred)
    tn, fp, fn, tp = cm.ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan
    return {
        'Accuracy':    accuracy_score(y_te, y_pred),
        'Precision':   precision_score(y_te, y_pred, zero_division=0),
        'Recall':      recall_score(y_te, y_pred, zero_division=0),
        'Specificity': specificity,
        'F1':          f1_score(y_te, y_pred, average='weighted'),
        'MCC':         matthews_corrcoef(y_te, y_pred),
        'ROC-AUC':     roc_auc_score(y_te, y_proba),
        'cm':          cm,
    }

# =============================================================================
# 11) EVALUATE ALL ML MODELS (fit on train, evaluate on untouched test)
# =============================================================================

def enhanced_evaluate(model, X_tr, y_tr, X_te, y_te, model_name):
    model.fit(X_tr, y_tr)
    y_pred = model.predict(X_te)
    y_proba = model.predict_proba(X_te)[:, 1]
    m = compute_full_metrics(y_te, y_pred, y_proba)

    print(f"\n{'='*50}\nMODEL: {model_name}\n{'='*50}")
    print(classification_report(y_te, y_pred))
    print(f"Accuracy: {m['Accuracy']:.4f} | Precision: {m['Precision']:.4f} | "
          f"Recall: {m['Recall']:.4f} | Specificity: {m['Specificity']:.4f} | "
          f"F1: {m['F1']:.4f} | MCC: {m['MCC']:.4f} | ROC-AUC: {m['ROC-AUC']:.4f}")
    return y_pred, y_proba, m

results = {}
fitted_pipelines = {}

for name, pipe in pipelines.items():
    y_pred, y_proba, m = enhanced_evaluate(pipe, X_train, y_train, X_test, y_test, name)
    m['y_pred'] = y_pred
    m['y_proba'] = y_proba
    results[name] = m
    fitted_pipelines[name] = pipe

# =============================================================================
# (A) FINAL RESULTS TABLE — Thermodynamic + 5 ML models, all metrics
#     -> This is the table you paste into the Results section as the
#        "classification report" summary table.
# =============================================================================

ALL_METRICS = ['Accuracy', 'Precision', 'Recall', 'Specificity', 'F1', 'MCC', 'ROC-AUC']

final_rows = [
    {'Model': 'Thermodynamic', **{k: metrics_thermo_naive[k] for k in ALL_METRICS}},
]
for name in MODEL_NAMES:
    final_rows.append({'Model': name, **{k: results[name][k] for k in ALL_METRICS}})

final_comparison = pd.DataFrame(final_rows)

print("\n" + "=" * 80)
print("(A) FINAL RESULTS TABLE — Thermodynamic vs ML MODELS — ALL METRICS")
print("=" * 80)
print(final_comparison.round(4).to_string(index=False))
final_comparison.round(4).to_csv('final_results_table.csv', index=False)

model_preds_for_excel = {'Thermodynamic': y_pred_thermo_naive}
for name in MODEL_NAMES:
    model_preds_for_excel[name] = results[name]['y_pred']
colors_map = {'Thermodynamic': THERMO_COLOR_NAIVE, **MODEL_COLOR}

export_classification_reports_excel(
    model_preds_for_excel, y_true, final_comparison,
    'classification_reports.xlsx', colors_map
)

def bar_color(model_name):
    if model_name == 'Thermodynamic':
        return THERMO_COLOR_NAIVE
    return MODEL_COLOR.get(model_name, '#7F7F7F')

# Best ML model by ROC-AUC (thermo excluded from this ranking)
best_ml_name = final_comparison[final_comparison['Model'] != 'Thermodynamic'] \
    .sort_values('ROC-AUC', ascending=False).iloc[0]['Model']
print(f"\nBest single ML model by ROC-AUC: {best_ml_name}")

# =============================================================================
# (A2) COMBINED METRIC COMPARISON CHART — AI models + Thermodynamic Model
#      side by side, across 4 key metrics (Specificity, MCC, ROC-AUC dropped
#      per author's request), 2x2 grid
# =============================================================================

A2_METRICS = ['Accuracy', 'Precision', 'Recall', 'F1']

fig = plt.figure(figsize=(10, 8))
gs = gridspec.GridSpec(2, 2, figure=fig, wspace=0.45, hspace=0.4)

for i, metric in enumerate(A2_METRICS):
    ax = fig.add_subplot(gs[i // 2, i % 2])
    data = final_comparison.sort_values(metric)
    colors = [bar_color(m) for m in data['Model']]
    ax.barh(data['Model'], data[metric], color=colors, edgecolor='white', linewidth=0.6, height=0.65)
    ax.set_title(f'({chr(97+i)}) {metric}', loc='left', fontweight='bold', fontsize=10)
    lo = min(0, data[metric].min() - 0.05)
    ax.set_xlim(lo, 1.05)
    ax.xaxis.set_major_locator(MaxNLocator(4))
    ax.tick_params(axis='y', labelsize=9)
    for idx, value in enumerate(data[metric]):
        ax.text(value + 0.015, idx, f"{value:.3f}", va='center', fontsize=7.5)

fig.suptitle('AI Models vs. Thermodynamic Model — Key Metrics Compared',
             y=1.0, fontsize=12, fontweight='bold')
plt.savefig('A2_ai_vs_thermo_all_metrics.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (A3) PER-CLASS PERFORMANCE CHART — Precision/Recall/F1 for Class 0
#      (No Scale) and Class 1 (Scale), all 6 models, one glance instead of a
#      long classification-report table
# =============================================================================

CLASS_LABELS = ['No Scale', 'Scale']
PERCLASS_METRICS = ['precision', 'recall', 'f1-score']
PERCLASS_METRIC_COLORS = {'precision': '#0072B2', 'recall': '#D55E00', 'f1-score': '#009E73'}

all_model_names_ordered = ['Thermodynamic'] + MODEL_NAMES

perclass_rows = []
for cls in CLASS_LABELS:
    for m in all_model_names_ordered:
        rep = classification_report(y_true, model_preds_for_excel[m],
                                     target_names=CLASS_LABELS, output_dict=True)
        for metric in PERCLASS_METRICS:
            perclass_rows.append({'Class': cls, 'Model': m, 'Metric': metric,
                                   'Value': rep[cls][metric]})
perclass_df = pd.DataFrame(perclass_rows)

fig, axes = plt.subplots(1, 2, figsize=(14, 6), sharey=True)

for ax, cls in zip(axes, CLASS_LABELS):
    sub = perclass_df[perclass_df['Class'] == cls]
    x = np.arange(len(all_model_names_ordered))
    width = 0.25
    for i, metric in enumerate(PERCLASS_METRICS):
        vals = [sub[(sub['Model'] == m) & (sub['Metric'] == metric)]['Value'].values[0]
                for m in all_model_names_ordered]
        bars = ax.bar(x + (i - 1) * width, vals, width,
                       label=metric.capitalize().replace('-score', '-score'),
                       color=PERCLASS_METRIC_COLORS[metric], edgecolor='white', linewidth=0.5)
        for b in bars:
            h = b.get_height()
            ax.text(b.get_x() + b.get_width() / 2, h + 0.015, f'{h:.2f}',
                     ha='center', va='bottom', fontsize=6.5, rotation=90)
    ax.set_xticks(x)
    ax.set_xticklabels(all_model_names_ordered, rotation=30, ha='right', fontsize=9)
    ax.set_title(f'Class: {cls}', fontweight='bold', fontsize=11)
    ax.set_ylim(0, 1.15)
    ax.yaxis.set_major_locator(MaxNLocator(6))

axes[0].set_ylabel('Score')
axes[0].legend(loc='upper left', ncol=3, fontsize=8, bbox_to_anchor=(0, 1.12))

fig.suptitle('Per-Class Performance — Precision / Recall / F1-score by Model',
             y=1.03, fontsize=12, fontweight='bold')
plt.savefig('A3_per_class_performance.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (B) COMBINED CONFUSION-MATRIX FIGURE — all 6 models, one panel each
# =============================================================================

cm_thermo_naive = confusion_matrix(y_true, y_pred_thermo_naive)

all_panels = [(cm_thermo_naive, 'Thermodynamic', THERMO_COLOR_NAIVE)] + \
             [(results[name]['cm'], name, MODEL_COLOR[name]) for name in MODEL_NAMES]

fig = plt.figure(figsize=(13, 8))
gs = gridspec.GridSpec(2, 3, figure=fig, wspace=0.4, hspace=0.5)

for i, (cm, title, color) in enumerate(all_panels):
    ax = fig.add_subplot(gs[i // 3, i % 3])
    cmap = sns.light_palette(color, as_cmap=True)
    sns.heatmap(cm, annot=True, fmt='d', cmap=cmap, cbar=False, square=True,
                annot_kws={'fontsize': 10, 'fontweight': 'bold'}, linewidths=0.5, linecolor='white',
                xticklabels=['No Scale', 'Scale'], yticklabels=['No Scale', 'Scale'], ax=ax)
    ax.set_title(f'({chr(97+i)}) {title}', loc='left', fontweight='bold', fontsize=9.5)
    ax.set_xlabel('Predicted', fontsize=8)
    ax.set_ylabel('Actual' if i % 3 == 0 else '', fontsize=8)
    ax.tick_params(length=0, labelsize=8)

fig.suptitle('Confusion Matrices — Thermodynamic Model vs. Machine-Learning Models',
             y=1.0, fontsize=12, fontweight='bold')
plt.savefig('B_confusion_matrices_combined.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (C1) RESAMPLING EFFECT — grouped bar chart, ADASYN vs No-ADASYN (5 ML models)
# =============================================================================

print("\n" + "=" * 70)
print("(C) ADASYN vs NO-ADASYN COMPARISON (5 ML models)")
print("=" * 70)

model_order = list(base_estimators.keys())
no_adasyn_results = {}     # store full metrics (incl. cm) per model, without ADASYN
no_adasyn_pipelines = {}   # store the fitted pipelines themselves (for saving/reload)

for name, est in base_estimators.items():
    pipe_no = build_pipeline(est, use_adasyn=False)
    pipe_no.fit(X_train, y_train)
    y_pred_no = pipe_no.predict(X_test)
    y_proba_no = pipe_no.predict_proba(X_test)[:, 1]
    m_no = compute_full_metrics(y_test, y_pred_no, y_proba_no)
    no_adasyn_results[name] = m_no
    no_adasyn_pipelines[name] = pipe_no

adasyn_rows = []
for name in model_order:
    adasyn_rows.append({'Model': f'{name} (ADASYN)', 'Accuracy': results[name]['Accuracy'],
                         'F1': results[name]['F1'], 'ROC-AUC': results[name]['ROC-AUC']})
    adasyn_rows.append({'Model': f'{name} (No ADASYN)', 'Accuracy': no_adasyn_results[name]['Accuracy'],
                         'F1': no_adasyn_results[name]['F1'], 'ROC-AUC': no_adasyn_results[name]['ROC-AUC']})

adasyn_comparison = pd.DataFrame(adasyn_rows)
print(adasyn_comparison.round(4).to_string(index=False))

metrics_ = ['Accuracy', 'F1', 'ROC-AUC']
fig = plt.figure(figsize=(11, 9))
gs = gridspec.GridSpec(3, 1, figure=fig, hspace=0.55, top=0.90)

for idx, metric in enumerate(metrics_):
    ax = fig.add_subplot(gs[idx, 0])
    x = np.arange(len(model_order))
    width = 0.36
    vals_ada = [adasyn_comparison.loc[adasyn_comparison['Model'] == f'{m} (ADASYN)', metric].values[0]
                for m in model_order]
    vals_no = [adasyn_comparison.loc[adasyn_comparison['Model'] == f'{m} (No ADASYN)', metric].values[0]
               for m in model_order]
    bars1 = ax.bar(x - width / 2, vals_ada, width, label='With ADASYN',
                    color='#2166AC', edgecolor='white', linewidth=0.6)
    bars2 = ax.bar(x + width / 2, vals_no, width, label='Without ADASYN',
                    color='#B2182B', edgecolor='white', linewidth=0.6)
    for bars in (bars1, bars2):
        for b in bars:
            h = b.get_height()
            ax.text(b.get_x() + b.get_width() / 2, h + 0.015, f'{h:.3f}',
                     ha='center', va='bottom', fontsize=7)
    ax.set_xticks(x)
    ax.set_xticklabels(model_order, fontsize=9)
    ax.set_ylabel(metric)
    ax.set_ylim(0, 1.08)
    ax.yaxis.set_major_locator(MaxNLocator(5))
    ax.set_title(f'({chr(97+idx)}) {metric}', loc='left', fontweight='bold', fontsize=10)
    if idx == 0:
        ax.legend(loc='lower center', bbox_to_anchor=(0.5, 1.18), ncol=2, fontsize=9, frameon=True)

fig.suptitle('Effect of ADASYN Resampling on ML Model Performance', y=0.995, fontsize=11, fontweight='bold')
plt.savefig('C1_adasyn_comparison.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (C2) CONFUSION-MATRIX COMPARISON — BEST model, with vs without ADASYN
# =============================================================================

cm_best_with = results[best_ml_name]['cm']
cm_best_without = no_adasyn_results[best_ml_name]['cm']

fig, axes = plt.subplots(1, 2, figsize=(9, 4.3))
for ax, cm, subtitle in zip(
        axes,
        [cm_best_with, cm_best_without],
        [f'{best_ml_name} — With ADASYN', f'{best_ml_name} — Without ADASYN']):
    cmap = sns.light_palette(MODEL_COLOR[best_ml_name], as_cmap=True)
    sns.heatmap(cm, annot=True, fmt='d', cmap=cmap, cbar=False, square=True,
                annot_kws={'fontsize': 12, 'fontweight': 'bold'}, linewidths=0.6, linecolor='white',
                xticklabels=['No Scale', 'Scale'], yticklabels=['No Scale', 'Scale'], ax=ax)
    ax.set_title(subtitle, fontweight='bold', fontsize=10)
    ax.set_xlabel('Predicted')
    ax.set_ylabel('Actual')
    ax.tick_params(length=0)

fig.suptitle('Effect of Resampling on the Best Model\'s Confusion Matrix', y=1.03,
             fontsize=11, fontweight='bold')
plt.savefig('C2_best_model_cm_adasyn_vs_none.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (D) COMBINED ROC-AUC CURVE — all 6 models
# =============================================================================

fig = plt.figure(figsize=(7, 6.5))
ax = fig.add_subplot(1, 1, 1)

fpr_n, tpr_n, _ = roc_curve(y_true, proba_thermo_naive)
ax.plot(fpr_n, tpr_n, color=THERMO_COLOR_NAIVE, linestyle=':', linewidth=1.8,
        label=f"Thermodynamic (AUC={metrics_thermo_naive['ROC-AUC']:.3f})")

for name in MODEL_NAMES:
    fpr, tpr, _ = roc_curve(y_test, results[name]['y_proba'])
    ax.plot(fpr, tpr, color=MODEL_COLOR[name], marker=MODEL_MARKER[name],
             markevery=0.1, markersize=5, linewidth=1.5,
             label=f"{name} (AUC={results[name]['ROC-AUC']:.3f})")

ax.plot([0, 1], [0, 1], color='0.6', linestyle='--', linewidth=1)
ax.set_xlabel('False Positive Rate')
ax.set_ylabel('True Positive Rate')
ax.set_title('ROC Curves — ML Models vs. Thermodynamic Model', fontweight='bold', fontsize=10)
ax.legend(loc='lower right', fontsize=8)
ax.set_xlim(0, 1)
ax.set_ylim(0, 1.02)
plt.savefig('D_roc_curves_all_models.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (E) PERMUTATION IMPORTANCE — 5 ML models, one combined figure
# =============================================================================

feature_names = X.columns.tolist()
perm_importances = {}

print("\nComputing permutation importance for all 5 ML models (test set)...")
for name in MODEL_NAMES:
    perm_result = permutation_importance(
        fitted_pipelines[name], X_test, y_test,
        n_repeats=20, random_state=42, scoring='roc_auc', n_jobs=-1
    )
    perm_df = pd.DataFrame({
        'Feature': feature_names,
        'Importance': perm_result.importances_mean,
        'Std': perm_result.importances_std,
    }).sort_values('Importance', ascending=False)
    perm_importances[name] = perm_df

fig = plt.figure(figsize=(14, 8))
gs = gridspec.GridSpec(2, 3, figure=fig, wspace=0.55, hspace=0.5)

for i, name in enumerate(MODEL_NAMES):
    ax = fig.add_subplot(gs[i // 3, i % 3])
    top = perm_importances[name].sort_values('Importance', ascending=True).tail(10)
    ax.barh(top['Feature'], top['Importance'], xerr=top['Std'],
            color=MODEL_COLOR[name], edgecolor='white', linewidth=0.6,
            error_kw=dict(elinewidth=0.8, capsize=2))
    ax.set_title(f'({chr(97+i)}) {name}', loc='left', fontweight='bold', fontsize=9.5)
    ax.set_xlabel('Permutation Importance (ROC-AUC drop)', fontsize=8)
    ax.tick_params(axis='y', labelsize=7.5)
    ax.xaxis.set_major_locator(MaxNLocator(4))

# hide the unused 6th panel (5 models in a 2x3 grid)
if len(MODEL_NAMES) < 6:
    ax_off = fig.add_subplot(gs[1, 2])
    ax_off.axis('off')

fig.suptitle('Permutation Importance — All ML Models (Test Set)', y=1.0, fontsize=12, fontweight='bold')
plt.savefig('E_permutation_importance_all_models.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (F) SHAP ANALYSIS — BEST ML model only, class = "Scale" only
# =============================================================================

print(f"\nRunning SHAP analysis for the best model: {best_ml_name}")

best_pipeline = fitted_pipelines[best_ml_name]
scaler_final = best_pipeline.named_steps['scaler']
clf_final = best_pipeline.named_steps['clf']

X_train_scaled = pd.DataFrame(scaler_final.transform(X_train), columns=feature_names)
X_test_scaled = pd.DataFrame(scaler_final.transform(X_test), columns=feature_names)

TREE_MODELS = ['Random Forest', 'Extra Trees', 'XGBoost', 'LightGBM']

if best_ml_name in TREE_MODELS:
    explainer = shap.TreeExplainer(clf_final)
    shap_values = explainer(X_test_scaled)
else:
    # MLP (or any non-tree model): use a generic Explainer with a background sample
    background = shap.sample(X_train_scaled, min(100, len(X_train_scaled)), random_state=42)
    explainer = shap.Explainer(clf_final.predict_proba, background)
    shap_values = explainer(X_test_scaled)

def extract_shap_for_class(sv, class_idx, feat_names):
    """
    Returns a shap.Explanation for a single class, regardless of whether the
    underlying explainer produced one set of values per class (ndim==3, most
    tree explainers on sklearn-style binary classifiers) or a single set of
    values assumed to represent the positive class (ndim==2, common for
    XGBoost/LightGBM TreeExplainer).
    """
    if hasattr(sv, "values") and sv.values.ndim == 3:
        vals = sv.values[:, :, class_idx]
        base = sv.base_values[:, class_idx] if np.ndim(sv.base_values) > 1 else sv.base_values
        return shap.Explanation(values=vals, base_values=base, data=sv.data, feature_names=feat_names)
    return sv  # single-output case already represents the positive ("Scale") class

sv_scale = extract_shap_for_class(shap_values, 1, feature_names)

plt.figure(figsize=(8, 7))
shap.summary_plot(sv_scale, X_test_scaled, show=False, plot_size=(8, 7))
plt.title(f'SHAP Summary — {best_ml_name} (Test Set, Class: Scale)', fontweight='bold', fontsize=11)
plt.savefig('F_shap_summary_best_model.png', dpi=300, bbox_inches='tight')
plt.show()

plt.figure(figsize=(7, 6))
shap.summary_plot(sv_scale, X_test_scaled, plot_type='bar', show=False, plot_size=(7, 6))
plt.title(f'SHAP Feature Importance — {best_ml_name} (Class: Scale)', fontweight='bold', fontsize=11)
plt.savefig('F_shap_bar_best_model.png', dpi=300, bbox_inches='tight')
plt.show()

# =============================================================================
# (G) SAVE ALL MODELS + DATA
#     -> lets you reload everything later (in this same script, see part H
#        below) and regenerate the exact same tables/figures WITHOUT
#        re-running Optuna tuning.
# =============================================================================

import joblib
import os

SAVE_DIR = 'saved_models'
os.makedirs(SAVE_DIR, exist_ok=True)

for name in MODEL_NAMES:
    joblib.dump(fitted_pipelines[name],
                os.path.join(SAVE_DIR, f"{name.replace(' ', '_')}_pipeline.joblib"))
    joblib.dump(no_adasyn_pipelines[name],
                os.path.join(SAVE_DIR, f"{name.replace(' ', '_')}_no_adasyn_pipeline.joblib"))

results_bundle = {
    'X_train':               X_train,
    'X_test':                X_test,
    'y_train':               y_train,
    'y_test':                y_test,
    'feature_names':         feature_names,
    'model_names':           MODEL_NAMES,
    'best_ml_name':          best_ml_name,
    'y_pred_thermo_naive':   y_pred_thermo_naive,
    'proba_thermo_naive':    proba_thermo_naive,
    'metrics_thermo_naive':  metrics_thermo_naive,
    'final_comparison':      final_comparison,
}
joblib.dump(results_bundle, os.path.join(SAVE_DIR, 'results_bundle.joblib'))

print(f"\nAll {len(MODEL_NAMES)} ML models (with & without ADASYN) and the results "
      f"bundle were saved to ./{SAVE_DIR}/")

print("\n" + "=" * 80)
print("PIPELINE COMPLETE (first pass — models trained, tuned, and saved)")
print("=" * 80)
print(final_comparison.round(4).to_string(index=False))

# =============================================================================
# (H) RELOAD THE SAVED MODELS FROM DISK AND RE-RUN EVERY RESULT/FIGURE ON THE
#     TEST DATA — ONE MORE TIME — using ONLY the loaded models (no Optuna, no
#     retraining). This proves the saved models reproduce the paper's figures
#     on their own and is the block you re-run whenever you just want to
#     regenerate figures (e.g. after a styling tweak).
# =============================================================================

print("\n" + "#" * 80)
print("# (H) RELOADING SAVED MODELS FROM DISK AND REPRODUCING ALL RESULTS")
print("#" * 80)

del fitted_pipelines, no_adasyn_pipelines, results, no_adasyn_results, final_comparison

loaded_bundle = joblib.load(os.path.join(SAVE_DIR, 'results_bundle.joblib'))
X_train              = loaded_bundle['X_train']
X_test               = loaded_bundle['X_test']
y_train              = loaded_bundle['y_train']
y_test               = loaded_bundle['y_test']
feature_names        = loaded_bundle['feature_names']
best_ml_name         = loaded_bundle['best_ml_name']
y_pred_thermo_naive  = loaded_bundle['y_pred_thermo_naive']
proba_thermo_naive   = loaded_bundle['proba_thermo_naive']
metrics_thermo_naive = loaded_bundle['metrics_thermo_naive']
y_true               = y_test.values

fitted_pipelines = {}
no_adasyn_pipelines = {}
for name in MODEL_NAMES:
    fitted_pipelines[name] = joblib.load(
        os.path.join(SAVE_DIR, f"{name.replace(' ', '_')}_pipeline.joblib"))
    no_adasyn_pipelines[name] = joblib.load(
        os.path.join(SAVE_DIR, f"{name.replace(' ', '_')}_no_adasyn_pipeline.joblib"))
print(f"Loaded {len(MODEL_NAMES)} ML models (with & without ADASYN) back from ./{SAVE_DIR}/")

# --- re-run predictions on the test set using ONLY the loaded models ---
results = {}
for name in MODEL_NAMES:
    pipe = fitted_pipelines[name]
    y_pred = pipe.predict(X_test)
    y_proba = pipe.predict_proba(X_test)[:, 1]
    m = compute_full_metrics(y_test, y_pred, y_proba)
    m['y_pred'] = y_pred
    m['y_proba'] = y_proba
    results[name] = m

no_adasyn_results = {}
for name in MODEL_NAMES:
    pipe_no = no_adasyn_pipelines[name]
    y_pred_no = pipe_no.predict(X_test)
    y_proba_no = pipe_no.predict_proba(X_test)[:, 1]
    no_adasyn_results[name] = compute_full_metrics(y_test, y_pred_no, y_proba_no)

# --- (A) rebuild the final results table ---
final_rows = [
    {'Model': 'Thermodynamic', **{k: metrics_thermo_naive[k] for k in ALL_METRICS}},
]
for name in MODEL_NAMES:
    final_rows.append({'Model': name, **{k: results[name][k] for k in ALL_METRICS}})
final_comparison = pd.DataFrame(final_rows)

print("\n" + "=" * 80)
print("(A, reloaded) FINAL RESULTS TABLE — Thermodynamic vs ML MODELS")
print("=" * 80)
print(final_comparison.round(4).to_string(index=False))
final_comparison.round(4).to_csv('final_results_table_reloaded.csv', index=False)

model_preds_for_excel_reloaded = {'Thermodynamic': y_pred_thermo_naive}
for name in MODEL_NAMES:
    model_preds_for_excel_reloaded[name] = results[name]['y_pred']

export_classification_reports_excel(
    model_preds_for_excel_reloaded, y_true, final_comparison,
    'classification_reports_reloaded.xlsx', colors_map
)

# --- (A2, reloaded) AI vs Thermodynamic combined metric chart ---
fig = plt.figure(figsize=(10, 8))
gs = gridspec.GridSpec(2, 2, figure=fig, wspace=0.45, hspace=0.4)
for i, metric in enumerate(A2_METRICS):
    ax = fig.add_subplot(gs[i // 2, i % 2])
    data = final_comparison.sort_values(metric)
    colors = [bar_color(m) for m in data['Model']]
    ax.barh(data['Model'], data[metric], color=colors, edgecolor='white', linewidth=0.6, height=0.65)
    ax.set_title(f'({chr(97+i)}) {metric}', loc='left', fontweight='bold', fontsize=10)
    lo = min(0, data[metric].min() - 0.05)
    ax.set_xlim(lo, 1.05)
    ax.xaxis.set_major_locator(MaxNLocator(4))
    ax.tick_params(axis='y', labelsize=9)
    for idx, value in enumerate(data[metric]):
        ax.text(value + 0.015, idx, f"{value:.3f}", va='center', fontsize=7.5)
fig.suptitle('[Reloaded] AI Models vs. Thermodynamic Model — Key Metrics Compared',
             y=1.0, fontsize=12, fontweight='bold')
plt.savefig('A2_ai_vs_thermo_all_metrics_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (A3, reloaded) per-class performance chart ---
perclass_rows = []
for cls in CLASS_LABELS:
    for m in all_model_names_ordered:
        rep = classification_report(y_true, model_preds_for_excel_reloaded[m],
                                     target_names=CLASS_LABELS, output_dict=True)
        for metric in PERCLASS_METRICS:
            perclass_rows.append({'Class': cls, 'Model': m, 'Metric': metric,
                                   'Value': rep[cls][metric]})
perclass_df = pd.DataFrame(perclass_rows)

fig, axes = plt.subplots(1, 2, figsize=(14, 6), sharey=True)
for ax, cls in zip(axes, CLASS_LABELS):
    sub = perclass_df[perclass_df['Class'] == cls]
    x = np.arange(len(all_model_names_ordered))
    width = 0.25
    for i, metric in enumerate(PERCLASS_METRICS):
        vals = [sub[(sub['Model'] == m) & (sub['Metric'] == metric)]['Value'].values[0]
                for m in all_model_names_ordered]
        bars = ax.bar(x + (i - 1) * width, vals, width,
                       label=metric.capitalize().replace('-score', '-score'),
                       color=PERCLASS_METRIC_COLORS[metric], edgecolor='white', linewidth=0.5)
        for b in bars:
            h = b.get_height()
            ax.text(b.get_x() + b.get_width() / 2, h + 0.015, f'{h:.2f}',
                     ha='center', va='bottom', fontsize=6.5, rotation=90)
    ax.set_xticks(x)
    ax.set_xticklabels(all_model_names_ordered, rotation=30, ha='right', fontsize=9)
    ax.set_title(f'Class: {cls}', fontweight='bold', fontsize=11)
    ax.set_ylim(0, 1.15)
    ax.yaxis.set_major_locator(MaxNLocator(6))
axes[0].set_ylabel('Score')
axes[0].legend(loc='upper left', ncol=3, fontsize=8, bbox_to_anchor=(0, 1.12))
fig.suptitle('[Reloaded] Per-Class Performance — Precision / Recall / F1-score by Model',
             y=1.03, fontsize=12, fontweight='bold')
plt.savefig('A3_per_class_performance_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (B, reloaded) combined confusion-matrix figure, all 6 models ---
cm_thermo_naive = confusion_matrix(y_true, y_pred_thermo_naive)
all_panels = [(cm_thermo_naive, 'Thermodynamic', THERMO_COLOR_NAIVE)] + \
             [(results[name]['cm'], name, MODEL_COLOR[name]) for name in MODEL_NAMES]

fig = plt.figure(figsize=(13, 8))
gs = gridspec.GridSpec(2, 3, figure=fig, wspace=0.4, hspace=0.5)
for i, (cm, title, color) in enumerate(all_panels):
    ax = fig.add_subplot(gs[i // 3, i % 3])
    cmap = sns.light_palette(color, as_cmap=True)
    sns.heatmap(cm, annot=True, fmt='d', cmap=cmap, cbar=False, square=True,
                annot_kws={'fontsize': 10, 'fontweight': 'bold'}, linewidths=0.5, linecolor='white',
                xticklabels=['No Scale', 'Scale'], yticklabels=['No Scale', 'Scale'], ax=ax)
    ax.set_title(f'({chr(97+i)}) {title}', loc='left', fontweight='bold', fontsize=9.5)
    ax.set_xlabel('Predicted', fontsize=8)
    ax.set_ylabel('Actual' if i % 3 == 0 else '', fontsize=8)
    ax.tick_params(length=0, labelsize=8)
fig.suptitle('[Reloaded] Confusion Matrices — Thermodynamic Model vs. ML Models',
             y=1.0, fontsize=12, fontweight='bold')
plt.savefig('B_confusion_matrices_combined_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (C1, reloaded) ADASYN vs No-ADASYN grouped bar chart ---
model_order = MODEL_NAMES
adasyn_rows = []
for name in model_order:
    adasyn_rows.append({'Model': f'{name} (ADASYN)', 'Accuracy': results[name]['Accuracy'],
                         'F1': results[name]['F1'], 'ROC-AUC': results[name]['ROC-AUC']})
    adasyn_rows.append({'Model': f'{name} (No ADASYN)', 'Accuracy': no_adasyn_results[name]['Accuracy'],
                         'F1': no_adasyn_results[name]['F1'], 'ROC-AUC': no_adasyn_results[name]['ROC-AUC']})
adasyn_comparison = pd.DataFrame(adasyn_rows)

metrics_ = ['Accuracy', 'F1', 'ROC-AUC']
fig = plt.figure(figsize=(11, 9))
gs = gridspec.GridSpec(3, 1, figure=fig, hspace=0.55, top=0.90)
for idx, metric in enumerate(metrics_):
    ax = fig.add_subplot(gs[idx, 0])
    x = np.arange(len(model_order))
    width = 0.36
    vals_ada = [adasyn_comparison.loc[adasyn_comparison['Model'] == f'{m} (ADASYN)', metric].values[0]
                for m in model_order]
    vals_no = [adasyn_comparison.loc[adasyn_comparison['Model'] == f'{m} (No ADASYN)', metric].values[0]
               for m in model_order]
    bars1 = ax.bar(x - width / 2, vals_ada, width, label='With ADASYN',
                    color='#2166AC', edgecolor='white', linewidth=0.6)
    bars2 = ax.bar(x + width / 2, vals_no, width, label='Without ADASYN',
                    color='#B2182B', edgecolor='white', linewidth=0.6)
    for bars in (bars1, bars2):
        for b in bars:
            h = b.get_height()
            ax.text(b.get_x() + b.get_width() / 2, h + 0.015, f'{h:.3f}',
                     ha='center', va='bottom', fontsize=7)
    ax.set_xticks(x)
    ax.set_xticklabels(model_order, fontsize=9)
    ax.set_ylabel(metric)
    ax.set_ylim(0, 1.08)
    ax.yaxis.set_major_locator(MaxNLocator(5))
    ax.set_title(f'({chr(97+idx)}) {metric}', loc='left', fontweight='bold', fontsize=10)
    if idx == 0:
        ax.legend(loc='lower center', bbox_to_anchor=(0.5, 1.18), ncol=2, fontsize=9, frameon=True)
fig.suptitle('[Reloaded] Effect of ADASYN Resampling on ML Model Performance',
             y=0.995, fontsize=11, fontweight='bold')
plt.savefig('C1_adasyn_comparison_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (C2, reloaded) confusion-matrix comparison for the best model ---
cm_best_with = results[best_ml_name]['cm']
cm_best_without = no_adasyn_results[best_ml_name]['cm']
fig, axes = plt.subplots(1, 2, figsize=(9, 4.3))
for ax, cm, subtitle in zip(
        axes,
        [cm_best_with, cm_best_without],
        [f'{best_ml_name} — With ADASYN', f'{best_ml_name} — Without ADASYN']):
    cmap = sns.light_palette(MODEL_COLOR[best_ml_name], as_cmap=True)
    sns.heatmap(cm, annot=True, fmt='d', cmap=cmap, cbar=False, square=True,
                annot_kws={'fontsize': 12, 'fontweight': 'bold'}, linewidths=0.6, linecolor='white',
                xticklabels=['No Scale', 'Scale'], yticklabels=['No Scale', 'Scale'], ax=ax)
    ax.set_title(subtitle, fontweight='bold', fontsize=10)
    ax.set_xlabel('Predicted')
    ax.set_ylabel('Actual')
    ax.tick_params(length=0)
fig.suptitle('[Reloaded] Effect of Resampling on the Best Model\'s Confusion Matrix',
             y=1.03, fontsize=11, fontweight='bold')
plt.savefig('C2_best_model_cm_adasyn_vs_none_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (D, reloaded) combined ROC-AUC curve, all 6 models ---
fig = plt.figure(figsize=(7, 6.5))
ax = fig.add_subplot(1, 1, 1)
fpr_n, tpr_n, _ = roc_curve(y_true, proba_thermo_naive)
ax.plot(fpr_n, tpr_n, color=THERMO_COLOR_NAIVE, linestyle=':', linewidth=1.8,
        label=f"Thermodynamic (AUC={metrics_thermo_naive['ROC-AUC']:.3f})")
for name in MODEL_NAMES:
    fpr, tpr, _ = roc_curve(y_test, results[name]['y_proba'])
    ax.plot(fpr, tpr, color=MODEL_COLOR[name], marker=MODEL_MARKER[name],
             markevery=0.1, markersize=5, linewidth=1.5,
             label=f"{name} (AUC={results[name]['ROC-AUC']:.3f})")
ax.plot([0, 1], [0, 1], color='0.6', linestyle='--', linewidth=1)
ax.set_xlabel('False Positive Rate')
ax.set_ylabel('True Positive Rate')
ax.set_title('[Reloaded] ROC Curves — ML Models vs. Thermodynamic Model', fontweight='bold', fontsize=10)
ax.legend(loc='lower right', fontsize=8)
ax.set_xlim(0, 1)
ax.set_ylim(0, 1.02)
plt.savefig('D_roc_curves_all_models_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (E, reloaded) permutation importance for the 5 ML models ---
perm_importances = {}
print("\n[Reloaded] Computing permutation importance for all 5 ML models (test set)...")
for name in MODEL_NAMES:
    perm_result = permutation_importance(
        fitted_pipelines[name], X_test, y_test,
        n_repeats=20, random_state=42, scoring='roc_auc', n_jobs=-1
    )
    perm_df = pd.DataFrame({
        'Feature': feature_names,
        'Importance': perm_result.importances_mean,
        'Std': perm_result.importances_std,
    }).sort_values('Importance', ascending=False)
    perm_importances[name] = perm_df

fig = plt.figure(figsize=(14, 8))
gs = gridspec.GridSpec(2, 3, figure=fig, wspace=0.55, hspace=0.5)
for i, name in enumerate(MODEL_NAMES):
    ax = fig.add_subplot(gs[i // 3, i % 3])
    top = perm_importances[name].sort_values('Importance', ascending=True).tail(10)
    ax.barh(top['Feature'], top['Importance'], xerr=top['Std'],
            color=MODEL_COLOR[name], edgecolor='white', linewidth=0.6,
            error_kw=dict(elinewidth=0.8, capsize=2))
    ax.set_title(f'({chr(97+i)}) {name}', loc='left', fontweight='bold', fontsize=9.5)
    ax.set_xlabel('Permutation Importance (ROC-AUC drop)', fontsize=8)
    ax.tick_params(axis='y', labelsize=7.5)
    ax.xaxis.set_major_locator(MaxNLocator(4))
if len(MODEL_NAMES) < 6:
    ax_off = fig.add_subplot(gs[1, 2])
    ax_off.axis('off')
fig.suptitle('[Reloaded] Permutation Importance — All ML Models (Test Set)',
             y=1.0, fontsize=12, fontweight='bold')
plt.savefig('E_permutation_importance_all_models_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

# --- (F, reloaded) SHAP analysis for the best ML model only, class = "Scale" only ---
print(f"\n[Reloaded] Running SHAP analysis for the best model: {best_ml_name}")
best_pipeline = fitted_pipelines[best_ml_name]
scaler_final = best_pipeline.named_steps['scaler']
clf_final = best_pipeline.named_steps['clf']
X_train_scaled = pd.DataFrame(scaler_final.transform(X_train), columns=feature_names)
X_test_scaled = pd.DataFrame(scaler_final.transform(X_test), columns=feature_names)

if best_ml_name in TREE_MODELS:
    explainer = shap.TreeExplainer(clf_final)
    shap_values = explainer(X_test_scaled)
else:
    background = shap.sample(X_train_scaled, min(100, len(X_train_scaled)), random_state=42)
    explainer = shap.Explainer(clf_final.predict_proba, background)
    shap_values = explainer(X_test_scaled)

sv_scale = extract_shap_for_class(shap_values, 1, feature_names)

plt.figure(figsize=(8, 7))
shap.summary_plot(sv_scale, X_test_scaled, show=False, plot_size=(8, 7))
plt.title(f'[Reloaded] SHAP Summary — {best_ml_name} (Test Set, Class: Scale)', fontweight='bold', fontsize=11)
plt.savefig('F_shap_summary_best_model_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

plt.figure(figsize=(7, 6))
shap.summary_plot(sv_scale, X_test_scaled, plot_type='bar', show=False, plot_size=(7, 6))
plt.title(f'[Reloaded] SHAP Feature Importance — {best_ml_name} (Class: Scale)', fontweight='bold', fontsize=11)
plt.savefig('F_shap_bar_best_model_reloaded.png', dpi=300, bbox_inches='tight')
plt.show()

print("\n" + "=" * 80)
print("PIPELINE COMPLETE — models trained once, saved, reloaded, and all")
print("results/figures reproduced on the test set from the reloaded models.")
print("=" * 80)
print(final_comparison.round(4).to_string(index=False))
